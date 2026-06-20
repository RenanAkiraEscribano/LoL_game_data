"""
run_experiments.py
==================
Orquestrador: 30 seeds × 4 tempos (10,15,20,25 min) × 8 métodos

Representação snapshot (último minuto <= T):
  - M1_LR_Snapshot  : Regressão Logística       — baseline linear
  - M1_RF_Snapshot  : Random Forest             — baseline não-linear clássico
  - M1_CB_Snapshot  : CatBoost                  — boosting com cat. nativas

Representação flat (todos os minutos 1..T concatenados):
  - M2_LR_VetorFlat : Regressão Logística       — baseline linear em alta dim
  - M2_RF_VetorFlat : Random Forest             — baseline não-linear clássico
  - M2_CB_VetorFlat : CatBoost                  — boosting com cat. nativas

Modelagem sequencial:
  - M3_GRU          : GRU causal + atenção temporal    — PyTorch
  - M4_Transformer  : Transformer Encoder + atenção    — PyTorch

Eixos de comparação:
  (a) Efeito da representação  : snapshot vs flat vs sequência
  (b) Efeito do modelo         : LR vs RF vs CatBoost vs GRU vs Transformer

Nota sobre M2_LR_VetorFlat:
  O vetor flat tem dimensão T×d (até ~1000 features para T=25).
  Usa LogisticRegressionCV com solver=saga (escala para alta dimensão)
  e seleção de C por cross-validation no treino, sem vazar val/teste.

Nota sobre M4_Transformer:
  Mesmos hiperparâmetros de treino do M3_GRU para isolamento arquitetural.
  Usa positional encoding sin/cos fixo + atenção temporal aprendível.
  Compartilha _Encoder e _SeqDS com o GRU (sem duplicação de código).

─── SPLITS COMPARTILHADOS ────────────────────────────────────────────────────
Para cada tempo T, todos os métodos operam sobre EXATAMENTE os mesmos
matchIds de treino, validação e teste. Isso é requisito para os testes
estatísticos pareados (Friedman + Wilcoxon): a diferença de desempenho
entre método A e B na seed s deve ser atribuível ao método, não à amostra.

Pipeline por T:
  1. build_common_ids(T)  → intersecção dos matchIds válidos em TODOS os loaders
  2. make_splits(ids, seed) → (tr_ids, va_ids, te_ids)  — estratificado, anti-leakage
  3. Cada runner filtra seu DataFrame/records pelos IDs recebidos

─── AUC ──────────────────────────────────────────────────────────────────────
Todos os métodos produzem probabilidades e reportam ROC-AUC.
CatBoost usa predict_proba() (disponível por padrão).

─── SAÍDA ────────────────────────────────────────────────────────────────────
  experiment_results/results_raw.csv      — linha por (método, tempo, seed)
  experiment_results/results_summary.xlsx — resumo + Friedman + Wilcoxon
  experiment_results/wilcoxon_<T>min.csv  — p-valores pareados por tempo

Retomada automática: seeds já concluídas são puladas.
"""

import os, glob, warnings, time
import numpy as np
import pandas as pd
from pathlib import Path
from itertools import combinations

warnings.filterwarnings("ignore")

# ══════════════════════════════════════════════════════════════════════════════
# CONFIG
# ══════════════════════════════════════════════════════════════════════════════
DATASET_DIR = r"riot_dump/BR1/challenger/ranked_solo_420/dataset_ts_csv"
MINUTES_LIST = [10, 15, 20, 25]
N_SEEDS = 30
SEEDS = list(range(N_SEEDS))

TEST_SIZE = 0.10
VAL_SIZE = 0.15
MIN_COVERAGE = 0.5  # fração mínima de minutos reais vs T (aplicada a todos)

# GRU — parâmetros sincronizados com mlGRU.py v3
BATCH_SIZE = 128
EPOCHS = 150
PATIENCE = 20
LR = 1e-3
HIDDEN_SIZE = 128
NUM_LAYERS = 2
DROPOUT_GRU = 0.1
DROPOUT_HEAD = 0.3
WEIGHT_DECAY = 1e-4
GRAD_CLIP = 1.0

# Transformer Encoder — hiperparâmetros
# D_MODEL < HIDDEN_SIZE para evitar overfit com ~8k amostras de treino.
# NHEAD deve dividir D_MODEL exatamente (64 / 4 = 16).
# Treino: idêntico ao GRU (EPOCHS, PATIENCE, LR, OneCycleLR, GRAD_CLIP)
# para que a comparação M3 vs M4 isole a arquitetura, não o protocolo.
D_MODEL = 64  # dimensão do modelo (projeção de entrada)
NHEAD = 4  # cabeças de atenção (D_MODEL / NHEAD = 16)
TF_LAYERS = 2  # camadas TransformerEncoderLayer
TF_DIM_FF = 128  # dimensão da FFN interna (2× D_MODEL)
TF_DROPOUT = 0.1  # dropout interno do Transformer

OUTPUT_DIR = Path("experiment_results")
OUTPUT_DIR.mkdir(exist_ok=True)
RAW_CSV = OUTPUT_DIR / "results_raw.csv"
SUMMARY_XLSX = OUTPUT_DIR / "results_summary.xlsx"

DROP_COLS = {
    "matchId",
    "game_datetime_utc",
    "gameCreation_ms",
    "y_blue_win",
    "patch",
    "gameDuration_s",
    "t_min",
    "wardsPlaced_Blue",
    "wardsPlaced_Red",
    "wardsKilled_Blue",
    "wardsKilled_Red",
    "pX_championName",
    "pX_posX",
    "pX_posY",
    "pX_wardsPlaced",
    "pX_wardsDestroyed"
}

# ── imports ───────────────────────────────────────────────────────────────────
try:
    from sklearn.model_selection import StratifiedShuffleSplit
    from sklearn.preprocessing import LabelEncoder, StandardScaler, OrdinalEncoder
    from sklearn.metrics import accuracy_score, f1_score, roc_auc_score
    from sklearn.linear_model import LogisticRegression, LogisticRegressionCV
    from sklearn.ensemble import RandomForestClassifier
    from sklearn.pipeline import Pipeline
    from sklearn.compose import ColumnTransformer
    from scipy.stats import friedmanchisquare, wilcoxon
    from catboost import CatBoostClassifier, Pool
    import torch, torch.nn as nn
    from torch.utils.data import Dataset, DataLoader

    DEVICE = torch.device("cuda" if torch.cuda.is_available() else "cpu")
    torch.backends.cudnn.deterministic = True
    torch.backends.cudnn.benchmark = False
except ImportError as e:
    raise SystemExit(f"Dependência faltando: {e}")


# ══════════════════════════════════════════════════════════════════════════════
# SPLITS COMPARTILHADOS
# ══════════════════════════════════════════════════════════════════════════════


def build_common_ids(dataset_dir: str, minute_cutoff: int) -> pd.DataFrame:
    """
    Retorna DataFrame com colunas [matchId, label] contendo apenas as partidas
    válidas em TODOS os loaders para este cutoff.

    Critérios (aplicados a cada CSV):
      - Colunas t_min e y_blue_win presentes
      - Ao menos 1 registro com t_min <= cutoff
      - t_min máximo observado >= int(cutoff * MIN_COVERAGE)
      - matchId único e não-nulo

    Ao usar a interseção, garantimos que snapshot, flat e GRU operam
    sobre o mesmo universo de partidas, tornando os splits pareáveis.
    """
    valid = {}
    for path in glob.glob(os.path.join(dataset_dir, "*.csv")):
        df = pd.read_csv(path)
        if "t_min" not in df.columns or "y_blue_win" not in df.columns:
            continue
        if "matchId" not in df.columns:
            continue
        mid = str(df["matchId"].iloc[0])
        if not mid or mid == "nan":
            continue
        df_t = df[df["t_min"] <= minute_cutoff]
        if df_t.empty:
            continue
        if df_t["t_min"].max() < int(minute_cutoff * MIN_COVERAGE):
            continue
        label = int(df["y_blue_win"].iloc[0])
        valid[mid] = label

    if not valid:
        raise RuntimeError(f"Nenhum matchId válido para T={minute_cutoff}.")

    out = (
        pd.DataFrame(
            [(mid, lbl) for mid, lbl in valid.items()], columns=["matchId", "label"]
        )
        .drop_duplicates("matchId")
        .reset_index(drop=True)
    )

    print(f"  T={minute_cutoff}: {len(out)} partidas comuns a todos os métodos")
    return out


def make_splits(id_df: pd.DataFrame, seed: int):
    """
    Divide id_df em (trainval, test) e (train, val) com estratificação.
    Retorna sets de matchIds: tr_ids, va_ids, te_ids.
    """
    y = id_df["label"].values
    sss_te = StratifiedShuffleSplit(1, test_size=TEST_SIZE, random_state=seed)
    tv_idx, te_idx = next(sss_te.split(range(len(id_df)), y))

    tv = id_df.iloc[tv_idx]
    te = id_df.iloc[te_idx]
    y_tv = tv["label"].values

    sss_va = StratifiedShuffleSplit(1, test_size=VAL_SIZE, random_state=seed + 1000)
    tr_idx, va_idx = next(sss_va.split(range(len(tv)), y_tv))

    tr = tv.iloc[tr_idx]
    va = tv.iloc[va_idx]

    tr_ids = set(tr["matchId"])
    va_ids = set(va["matchId"])
    te_ids = set(te["matchId"])

    # Verificação anti-leakage
    assert not (tr_ids & te_ids), f"Leakage tr/te seed={seed}"
    assert not (va_ids & te_ids), f"Leakage va/te seed={seed}"
    assert not (tr_ids & va_ids), f"Leakage tr/va seed={seed}"

    return tr_ids, va_ids, te_ids


# ══════════════════════════════════════════════════════════════════════════════
# LOADERS — retornam estruturas filtradas por matchId
# ══════════════════════════════════════════════════════════════════════════════


def load_snapshot(dataset_dir: str, minute_cutoff: int, valid_ids: set) -> pd.DataFrame:
    """Snapshot: última linha com t_min <= cutoff, filtrado por valid_ids."""
    rows = []
    for path in glob.glob(os.path.join(dataset_dir, "*.csv")):
        df = pd.read_csv(path)
        if "t_min" not in df.columns or "y_blue_win" not in df.columns:
            continue
        if "matchId" not in df.columns:
            continue
        mid = str(df["matchId"].iloc[0])
        if mid not in valid_ids:
            continue
        eligible = df[df["t_min"] <= minute_cutoff].sort_values("t_min")
        if eligible.empty:
            continue
        rows.append(eligible.iloc[-1])
    if not rows:
        raise RuntimeError("load_snapshot: nenhuma partida carregada.")
    out = pd.DataFrame(rows)
    return out.drop_duplicates("matchId").reset_index(drop=True)


def load_flat(dataset_dir: str, minute_cutoff: int, valid_ids: set) -> pd.DataFrame:
    """Flat: minutos 1..T concatenados como colunas, filtrado por valid_ids."""
    rows = []
    for path in glob.glob(os.path.join(dataset_dir, "*.csv")):
        df = pd.read_csv(path)
        if "t_min" not in df.columns or "y_blue_win" not in df.columns:
            continue
        if "matchId" not in df.columns:
            continue
        mid = str(df["matchId"].iloc[0])
        if mid not in valid_ids:
            continue
        df = df.sort_values("t_min").reset_index(drop=True)
        df = df[df["t_min"] <= minute_cutoff]
        if df.empty:
            continue
        feat_cols = [c for c in df.columns if c not in DROP_COLS]
        midx = pd.RangeIndex(1, minute_cutoff + 1, name="t_min")
        reindexed = (
            df.groupby("t_min")[feat_cols]
            .last()
            .reindex(midx)
            .ffill()
            .bfill()
            .fillna(0)
        )
        flat = {
            f"{col}__t{t}": reindexed.at[t, col]
            for t in range(1, minute_cutoff + 1)
            for col in feat_cols
        }
        flat["matchId"] = mid
        flat["y_blue_win"] = int(df["y_blue_win"].iloc[0])
        rows.append(flat)
    if not rows:
        raise RuntimeError("load_flat: nenhuma partida carregada.")
    out = pd.DataFrame(rows)
    return out.drop_duplicates("matchId").reset_index(drop=True)


def load_gru(dataset_dir: str, minute_cutoff: int, valid_ids: set):
    """Sequências 3D para a GRU, filtradas por valid_ids."""
    records = []
    for path in glob.glob(os.path.join(dataset_dir, "*.csv")):
        df = pd.read_csv(path)
        if df.empty or "t_min" not in df.columns or "y_blue_win" not in df.columns:
            continue
        if "matchId" not in df.columns:
            continue
        mid = str(df["matchId"].iloc[0])
        if mid not in valid_ids:
            continue
        df = df.sort_values("t_min").reset_index(drop=True)
        df = df[df["t_min"] <= minute_cutoff]
        if df.empty:
            continue
        feat_cols = [c for c in df.columns if c not in DROP_COLS]
        midx = pd.RangeIndex(1, minute_cutoff + 1, name="t_min")
        seq_df = (
            df.groupby("t_min")[feat_cols]
            .last()
            .reindex(midx)
            .ffill()
            .bfill()
            .fillna(0)
        )
        records.append(
            {
                "match_id": mid,
                "label": int(df["y_blue_win"].iloc[0]),
                "seq_df": seq_df,
                "feat_cols": feat_cols,
            }
        )
    if not records:
        raise RuntimeError("load_gru: nenhuma partida carregada.")

    # Interseção de colunas para uniformidade entre registros
    common_cols = set(records[0]["feat_cols"])
    for r in records[1:]:
        common_cols &= set(r["feat_cols"])
    feat_cols = [c for c in records[0]["feat_cols"] if c in common_cols]
    cat_cols = [
        c for c in feat_cols if "_championName" in c or "_individualPosition" in c
    ]
    num_cols = [c for c in feat_cols if c not in cat_cols]
    for r in records:
        r["seq_df"] = r["seq_df"][[c for c in feat_cols if c in r["seq_df"].columns]]

    return records, cat_cols, num_cols


# ══════════════════════════════════════════════════════════════════════════════
# HELPERS DE PREPARAÇÃO
# ══════════════════════════════════════════════════════════════════════════════


def _filter_by_ids(df: pd.DataFrame, ids: set) -> pd.DataFrame:
    """Filtra DataFrame mantendo apenas linhas cujo matchId está em ids."""
    return df[df["matchId"].isin(ids)].reset_index(drop=True)


def _filter_records(records: list, ids: set) -> list:
    """Filtra lista de records da GRU pelo conjunto de matchIds."""
    return [r for r in records if r["match_id"] in ids]


def prepare_snapshot_cb(df):
    """Features para CatBoost: separa cat explicitamente."""
    y = df["y_blue_win"].astype(int)
    X = df.drop(
        columns=[c for c in df.columns if c in DROP_COLS], errors="ignore"
    ).copy()
    cat = [
        c
        for c in X.columns
        if c.endswith("_championName") or c.endswith("_individualPosition")
    ]
    for c in cat:
        X[c] = X[c].fillna("").astype(str)
    num = [c for c in X.columns if c not in cat]
    X[num] = X[num].apply(pd.to_numeric, errors="coerce").fillna(0)
    return X, y, cat


def prepare_snapshot_sklearn(df):
    """Features para LR/RF: num + cat separados para Pipeline."""
    y = df["y_blue_win"].astype(int)
    X = df.drop(
        columns=[c for c in df.columns if c in DROP_COLS], errors="ignore"
    ).copy()
    cat = [
        c
        for c in X.columns
        if c.endswith("_championName") or c.endswith("_individualPosition")
    ]
    num = [c for c in X.columns if c not in cat]
    for c in cat:
        X[c] = X[c].fillna("__unknown__").astype(str)
    X[num] = X[num].apply(pd.to_numeric, errors="coerce").fillna(0)
    return X, y, num, cat


def prepare_flat_cb(df):
    """Features flat para CatBoost."""
    y = df["y_blue_win"].astype(int)
    X = df.drop(columns=["matchId", "y_blue_win"], errors="ignore").copy()
    cat = [
        c for c in X.columns if "_championName__t" in c or "_individualPosition__t" in c
    ]
    for c in cat:
        X[c] = X[c].fillna("").astype(str)
    num = [c for c in X.columns if c not in cat]
    X[num] = X[num].apply(pd.to_numeric, errors="coerce").fillna(0)
    return X, y, cat


def prepare_flat_sklearn(df):
    """Features flat para LR/RF."""
    y = df["y_blue_win"].astype(int)
    X = df.drop(columns=["matchId", "y_blue_win"], errors="ignore").copy()
    cat = [
        c for c in X.columns if "_championName__t" in c or "_individualPosition__t" in c
    ]
    num = [c for c in X.columns if c not in cat]
    for c in cat:
        X[c] = X[c].fillna("__unknown__").astype(str)
    X[num] = X[num].apply(pd.to_numeric, errors="coerce").fillna(0)
    return X, y, num, cat


def _sklearn_pipe(seed, num_cols, cat_cols, clf):
    """Pipeline padrão: StandardScaler (num) + OrdinalEncoder (cat) + clf."""
    pre = ColumnTransformer(
        [
            ("num", StandardScaler(), num_cols),
            (
                "cat",
                OrdinalEncoder(handle_unknown="use_encoded_value", unknown_value=-1),
                cat_cols,
            ),
        ],
        remainder="drop",
    )
    return Pipeline([("pre", pre), ("clf", clf)])


def _result(yte, yp, yprob, n_tr, n_va, n_te, train_s, inf_s, **extra):
    return {
        "accuracy": accuracy_score(yte, yp),
        "f1": f1_score(yte, yp),
        "auc": roc_auc_score(yte, yprob) if len(set(yte)) > 1 else float("nan"),
        "n_train": n_tr,
        "n_val": n_va,
        "n_test": n_te,
        "train_s": round(train_s, 4),
        "inf_s": round(inf_s, 6),
        **extra,
    }


# ══════════════════════════════════════════════════════════════════════════════
# RUNNERS — recebem dados pré-carregados + splits por matchId
# ══════════════════════════════════════════════════════════════════════════════


def run_lr_snapshot_seeds(snap_df, splits_by_seed, seeds):
    """M1_LR_Snapshot — Regressão Logística no snapshot (baseline linear)."""
    results = {}
    for seed in seeds:
        tr_ids, va_ids, te_ids = splits_by_seed[seed]
        tr = _filter_by_ids(snap_df, tr_ids)
        va = _filter_by_ids(snap_df, va_ids)
        te = _filter_by_ids(snap_df, te_ids)

        Xtr, ytr, num_c, cat_c = prepare_snapshot_sklearn(tr)
        _, yva, _, _ = prepare_snapshot_sklearn(va)
        Xte, yte, _, _ = prepare_snapshot_sklearn(te)

        pipe = _sklearn_pipe(
            seed,
            num_c,
            cat_c,
            LogisticRegression(
                max_iter=1000, solver="lbfgs", C=1.0, random_state=seed, n_jobs=-1
            ),
        )
        t0 = time.perf_counter()
        pipe.fit(Xtr, ytr)
        train_s = time.perf_counter() - t0
        t1 = time.perf_counter()
        yp = pipe.predict(Xte)
        yprob = pipe.predict_proba(Xte)[:, 1]
        inf_s = time.perf_counter() - t1

        results[seed] = _result(
            yte, yp, yprob, len(ytr), len(yva), len(yte), train_s, inf_s
        )
    return results


def run_lr_flat_seeds(flat_df, splits_by_seed, seeds):
    """M2_LR_VetorFlat — Regressão Logística no vetor achatado.

    Usa LogisticRegressionCV com solver=saga (escala para T×d ≈ 1000 features)
    e seleção de C por cross-validation (cv=3) no conjunto de treino.
    Isso torna a regularização adaptativa à dimensionalidade de cada cutoff T,
    sem vazar informação de validação ou teste.
    """
    results = {}
    for seed in seeds:
        tr_ids, va_ids, te_ids = splits_by_seed[seed]
        tr = _filter_by_ids(flat_df, tr_ids)
        va = _filter_by_ids(flat_df, va_ids)
        te = _filter_by_ids(flat_df, te_ids)

        Xtr, ytr, num_c, cat_c = prepare_flat_sklearn(tr)
        _, yva, _, _ = prepare_flat_sklearn(va)
        Xte, yte, _, _ = prepare_flat_sklearn(te)

        # Pipeline: pré-processamento + LR com CV de regularização
        pipe = _sklearn_pipe(
            seed,
            num_c,
            cat_c,
            LogisticRegressionCV(
                Cs=[0.001, 0.01, 0.1, 1.0, 10.0],
                cv=3,
                solver="saga",  # único solver que escala para alta dim
                max_iter=3000,  # saga precisa de mais iterações que lbfgs
                n_jobs=-1,
                random_state=seed,
            ),
        )
        t0 = time.perf_counter()
        pipe.fit(Xtr, ytr)
        train_s = time.perf_counter() - t0

        t1 = time.perf_counter()
        yp = pipe.predict(Xte)
        yprob = pipe.predict_proba(Xte)[:, 1]
        inf_s = time.perf_counter() - t1

        results[seed] = _result(
            yte, yp, yprob, len(ytr), len(yva), len(yte), train_s, inf_s
        )
    return results


def run_rf_snapshot_seeds(snap_df, splits_by_seed, seeds):
    """M1_RF_Snapshot."""
    results = {}
    for seed in seeds:
        tr_ids, va_ids, te_ids = splits_by_seed[seed]
        tr = _filter_by_ids(snap_df, tr_ids)
        va = _filter_by_ids(snap_df, va_ids)
        te = _filter_by_ids(snap_df, te_ids)

        Xtr, ytr, num_c, cat_c = prepare_snapshot_sklearn(tr)
        _, yva, _, _ = prepare_snapshot_sklearn(va)
        Xte, yte, _, _ = prepare_snapshot_sklearn(te)

        pipe = _sklearn_pipe(
            seed,
            num_c,
            cat_c,
            RandomForestClassifier(
                n_estimators=300, max_features="sqrt", random_state=seed, n_jobs=-1
            ),
        )
        t0 = time.perf_counter()
        pipe.fit(Xtr, ytr)
        train_s = time.perf_counter() - t0
        t1 = time.perf_counter()
        yp = pipe.predict(Xte)
        yprob = pipe.predict_proba(Xte)[:, 1]
        inf_s = time.perf_counter() - t1

        results[seed] = _result(
            yte, yp, yprob, len(ytr), len(yva), len(yte), train_s, inf_s
        )
    return results


def run_cb_snapshot_seeds(snap_df, splits_by_seed, seeds):
    """M1_CB_Snapshot — CatBoost com predict_proba para AUC."""
    results = {}
    for seed in seeds:
        tr_ids, va_ids, te_ids = splits_by_seed[seed]
        tr = _filter_by_ids(snap_df, tr_ids)
        va = _filter_by_ids(snap_df, va_ids)
        te = _filter_by_ids(snap_df, te_ids)

        Xtr, ytr, cat = prepare_snapshot_cb(tr)
        Xva, yva, _ = prepare_snapshot_cb(va)
        Xte, yte, _ = prepare_snapshot_cb(te)

        model = CatBoostClassifier(
            random_seed=seed, verbose=0, early_stopping_rounds=50
        )
        t0 = time.perf_counter()
        model.fit(
            Pool(Xtr, ytr, cat_features=cat),
            eval_set=Pool(Xva, yva, cat_features=cat),
            use_best_model=True,
            verbose=False,
        )
        train_s = time.perf_counter() - t0

        t1 = time.perf_counter()
        yp = model.predict(Pool(Xte, yte, cat_features=cat)).astype(int).reshape(-1)
        yprob = model.predict_proba(Pool(Xte, yte, cat_features=cat))[:, 1]
        inf_s = time.perf_counter() - t1

        results[seed] = _result(
            yte, yp, yprob, len(ytr), len(yva), len(yte), train_s, inf_s
        )
    return results


def run_rf_flat_seeds(flat_df, splits_by_seed, seeds):
    """M2_RF_VetorFlat."""
    results = {}
    for seed in seeds:
        tr_ids, va_ids, te_ids = splits_by_seed[seed]
        tr = _filter_by_ids(flat_df, tr_ids)
        va = _filter_by_ids(flat_df, va_ids)
        te = _filter_by_ids(flat_df, te_ids)

        Xtr, ytr, num_c, cat_c = prepare_flat_sklearn(tr)
        _, yva, _, _ = prepare_flat_sklearn(va)
        Xte, yte, _, _ = prepare_flat_sklearn(te)

        pipe = _sklearn_pipe(
            seed,
            num_c,
            cat_c,
            RandomForestClassifier(
                n_estimators=300, max_features="sqrt", random_state=seed, n_jobs=-1
            ),
        )
        t0 = time.perf_counter()
        pipe.fit(Xtr, ytr)
        train_s = time.perf_counter() - t0
        t1 = time.perf_counter()
        yp = pipe.predict(Xte)
        yprob = pipe.predict_proba(Xte)[:, 1]
        inf_s = time.perf_counter() - t1

        results[seed] = _result(
            yte, yp, yprob, len(ytr), len(yva), len(yte), train_s, inf_s
        )
    return results


def run_cb_flat_seeds(flat_df, splits_by_seed, seeds):
    """M2_CB_VetorFlat — CatBoost com predict_proba para AUC."""
    results = {}
    for seed in seeds:
        tr_ids, va_ids, te_ids = splits_by_seed[seed]
        tr = _filter_by_ids(flat_df, tr_ids)
        va = _filter_by_ids(flat_df, va_ids)
        te = _filter_by_ids(flat_df, te_ids)

        Xtr, ytr, cat = prepare_flat_cb(tr)
        Xva, yva, _ = prepare_flat_cb(va)
        Xte, yte, _ = prepare_flat_cb(te)

        model = CatBoostClassifier(
            random_seed=seed, verbose=0, early_stopping_rounds=50
        )
        t0 = time.perf_counter()
        model.fit(
            Pool(Xtr, ytr, cat_features=cat),
            eval_set=Pool(Xva, yva, cat_features=cat),
            use_best_model=True,
            verbose=False,
        )
        train_s = time.perf_counter() - t0

        t1 = time.perf_counter()
        yp = model.predict(Pool(Xte, yte, cat_features=cat)).astype(int).reshape(-1)
        yprob = model.predict_proba(Pool(Xte, yte, cat_features=cat))[:, 1]
        inf_s = time.perf_counter() - t1

        results[seed] = _result(
            yte, yp, yprob, len(ytr), len(yva), len(yte), train_s, inf_s
        )
    return results


# ══════════════════════════════════════════════════════════════════════════════
# GRU — encoder, dataset, modelo, treino/avaliação
# ══════════════════════════════════════════════════════════════════════════════


def _dynamic_emb_dim(vocab_size: int) -> int:
    return min(50, (vocab_size + 1) // 2)


class _Encoder:
    def __init__(self, cat_cols, num_cols):
        self.cat_cols = cat_cols
        self.num_cols = num_cols
        self.les = {c: LabelEncoder() for c in cat_cols}
        self.scaler = StandardScaler()

    @property
    def cat_vocab_sizes(self):
        return [len(self.les[c].classes_) for c in self.cat_cols]

    @property
    def cat_emb_dims(self):
        return [_dynamic_emb_dim(v) for v in self.cat_vocab_sizes]

    def fit(self, records):
        cat_data = {c: [] for c in self.cat_cols}
        num_rows = []
        for r in records:
            df = r["seq_df"]
            for c in self.cat_cols:
                cat_data[c].extend(df[c].fillna("").astype(str).tolist())
            if self.num_cols:
                num_rows.append(df[self.num_cols].values)
        for c in self.cat_cols:
            self.les[c].fit(cat_data[c] + ["__unknown__"])
        if num_rows:
            self.scaler.fit(np.vstack(num_rows))

    def transform(self, rec):
        df = rec["seq_df"].copy()
        if self.num_cols:
            x_num = self.scaler.transform(
                df[self.num_cols].values.astype(float)
            ).astype(np.float32)
        else:
            x_num = np.zeros((len(df), 0), dtype=np.float32)

        cat_encoded = []
        for c in self.cat_cols:
            vals = (
                df[c].fillna("").astype(str).tolist()
                if c in df.columns
                else ["__unknown__"] * len(df)
            )
            enc = [
                (
                    self.les[c].transform([v])[0]
                    if v in self.les[c].classes_
                    else self.les[c].transform(["__unknown__"])[0]
                )
                for v in vals
            ]
            cat_encoded.append(np.array(enc, dtype=np.int64).reshape(-1, 1))
        x_cat = (
            np.hstack(cat_encoded)
            if cat_encoded
            else np.zeros((len(df), 0), dtype=np.int64)
        )
        return x_num, x_cat


class _SeqDS(Dataset):
    def __init__(self, records, encoder):
        self.x_nums, self.x_cats, self.y = [], [], []
        for r in records:
            x_num, x_cat = encoder.transform(r)
            self.x_nums.append(torch.tensor(x_num, dtype=torch.float32))
            self.x_cats.append(torch.tensor(x_cat, dtype=torch.long))
            self.y.append(r["label"])

    def __len__(self):
        return len(self.y)

    def __getitem__(self, i):
        return (
            self.x_nums[i],
            self.x_cats[i],
            torch.tensor(self.y[i], dtype=torch.float32),
        )


class _GRU(nn.Module):
    """GRU causal + atenção temporal + embeddings dinâmicos (mlGRU.py v3)."""

    def __init__(self, num_size, cat_vocab_sizes, cat_emb_dims):
        super().__init__()
        self.embeddings = nn.ModuleList(
            [
                nn.Embedding(vocab, emb)
                for vocab, emb in zip(cat_vocab_sizes, cat_emb_dims)
            ]
        )
        gru_input = num_size + sum(cat_emb_dims)
        self.gru = nn.GRU(
            gru_input,
            HIDDEN_SIZE,
            NUM_LAYERS,
            batch_first=True,
            dropout=DROPOUT_GRU if NUM_LAYERS > 1 else 0.0,
            bidirectional=False,
        )
        self.attn = nn.Linear(HIDDEN_SIZE, 1)
        self.head = nn.Sequential(
            nn.LayerNorm(HIDDEN_SIZE),
            nn.Linear(HIDDEN_SIZE, 32),
            nn.ReLU(),
            nn.Dropout(DROPOUT_HEAD),
            nn.Linear(32, 1),
        )

    def forward(self, x_num, x_cat):
        embs = [e(x_cat[..., i]) for i, e in enumerate(self.embeddings)]
        x = torch.cat([x_num] + embs, dim=-1) if embs else x_num
        out, _ = self.gru(x)
        w = torch.softmax(self.attn(out), dim=1)
        return self.head((w * out).sum(dim=1)).squeeze(1)


def _seq_train(model, loader, opt, crit, sched):
    model.train()
    tot = 0.0
    for xn, xc, y in loader:
        xn, xc, y = xn.to(DEVICE), xc.to(DEVICE), y.to(DEVICE)
        opt.zero_grad()
        loss = crit(model(xn, xc), y)
        loss.backward()
        nn.utils.clip_grad_norm_(model.parameters(), GRAD_CLIP)
        opt.step()
        sched.step()
        tot += loss.item() * len(y)
    return tot / len(loader.dataset)


@torch.no_grad()
def _seq_eval(model, loader, crit):
    model.eval()
    tot, preds, probs, trues = 0.0, [], [], []
    for xn, xc, y in loader:
        xn, xc, y = xn.to(DEVICE), xc.to(DEVICE), y.to(DEVICE)
        logits = model(xn, xc)
        tot += crit(logits, y).item() * len(y)
        prob = torch.sigmoid(logits).cpu()
        probs += prob.tolist()
        preds += (prob >= 0.5).int().tolist()
        trues += y.cpu().int().tolist()
    return tot / len(loader.dataset), preds, probs, trues


def run_gru_seeds(all_records, cat_cols, num_cols, splits_by_seed, seeds):
    """M3_GRU."""
    results = {}
    for seed in seeds:
        torch.manual_seed(seed)
        np.random.seed(seed)
        tr_ids, va_ids, te_ids = splits_by_seed[seed]
        tr = _filter_records(all_records, tr_ids)
        va = _filter_records(all_records, va_ids)
        te = _filter_records(all_records, te_ids)

        enc = _Encoder(cat_cols, num_cols)
        enc.fit(tr)

        g = torch.Generator()
        g.manual_seed(seed)
        _pin = DEVICE.type == "cuda"
        _nw = min(4, os.cpu_count() or 1) if _pin else 0
        mk = lambda recs, sh: DataLoader(
            _SeqDS(recs, enc),
            BATCH_SIZE,
            shuffle=sh,
            generator=g if sh else None,
            pin_memory=_pin,
            num_workers=_nw,
            persistent_workers=(_nw > 0),
        )
        tr_l, va_l, te_l = mk(tr, True), mk(va, False), mk(te, False)

        n_neg = sum(1 for r in tr if r["label"] == 0)
        n_pos = sum(1 for r in tr if r["label"] == 1)
        pos_w = torch.tensor([n_neg / n_pos], dtype=torch.float32).to(DEVICE)

        model = _GRU(len(num_cols), enc.cat_vocab_sizes, enc.cat_emb_dims).to(DEVICE)
        opt = torch.optim.Adam(model.parameters(), lr=LR, weight_decay=WEIGHT_DECAY)
        crit = nn.BCEWithLogitsLoss(pos_weight=pos_w)
        sched = torch.optim.lr_scheduler.OneCycleLR(
            opt,
            max_lr=LR,
            steps_per_epoch=len(tr_l),
            epochs=EPOCHS,
            pct_start=0.1,
            anneal_strategy="cos",
        )

        best_loss, best_state, wait, epochs_run = float("inf"), None, 0, 0
        t0 = time.perf_counter()
        for _ in range(EPOCHS):
            epochs_run += 1
            _seq_train(model, tr_l, opt, crit, sched)
            vl, _, _, _ = _seq_eval(model, va_l, crit)
            if vl < best_loss:
                best_loss = vl
                best_state = {k: v.cpu().clone() for k, v in model.state_dict().items()}
                wait = 0
            else:
                wait += 1
                if wait >= PATIENCE:
                    break
        train_s = time.perf_counter() - t0

        model.load_state_dict(best_state)
        model.to(DEVICE)
        t1 = time.perf_counter()
        _, yp, yprob, yt = _seq_eval(model, te_l, crit)
        inf_s = time.perf_counter() - t1

        results[seed] = _result(
            yt,
            yp,
            yprob,
            len(tr),
            len(va),
            len(te),
            train_s,
            inf_s,
            epochs_run=epochs_run,
        )
    return results


# ══════════════════════════════════════════════════════════════════════════════
# TRANSFORMER — modelo, treino/avaliação
# ══════════════════════════════════════════════════════════════════════════════


class _PositionalEncoding(nn.Module):
    """Positional encoding sin/cos fixo (Vaswani et al. 2017).

    Fixo (não aprendível) para evitar overfitting adicional
    em datasets de tamanho moderado (~8k amostras de treino).
    Sequências curtas (T=10..25) já fornecem sinal posicional suficiente.
    """

    def __init__(self, d_model: int, max_len: int = 50, dropout: float = 0.1):
        super().__init__()
        self.dropout = nn.Dropout(dropout)
        pe = torch.zeros(max_len, d_model)
        pos = torch.arange(0, max_len, dtype=torch.float).unsqueeze(1)
        div = torch.exp(
            torch.arange(0, d_model, 2, dtype=torch.float)
            * (-torch.log(torch.tensor(10000.0)) / d_model)
        )
        pe[:, 0::2] = torch.sin(pos * div)
        pe[:, 1::2] = torch.cos(pos * div)
        self.register_buffer("pe", pe.unsqueeze(0))  # (1, max_len, d_model)

    def forward(self, x):
        # x: (batch, T, d_model)
        return self.dropout(x + self.pe[:, : x.size(1)])


class _Transformer(nn.Module):
    """Transformer Encoder causal para predição de resultado de partida.

    Arquitetura:
      embeddings cat + features num
        → projeção linear → D_MODEL
        → PositionalEncoding (sin/cos fixo)
        → TransformerEncoderLayer × TF_LAYERS
        → atenção temporal aprendível (softmax sobre T)
        → cabeça de classificação idêntica à GRU

    Compartilha _Encoder e _SeqDS com o GRU.
    Hiperparâmetros de treino idênticos ao GRU para isolar a arquitetura.
    """

    def __init__(self, num_size: int, cat_vocab_sizes: list, cat_emb_dims: list):
        super().__init__()

        # Embeddings categóricos (idênticos ao GRU)
        self.embeddings = nn.ModuleList(
            [
                nn.Embedding(vocab, emb)
                for vocab, emb in zip(cat_vocab_sizes, cat_emb_dims)
            ]
        )
        input_dim = num_size + sum(cat_emb_dims)

        # Projeção de entrada → D_MODEL
        self.input_proj = nn.Linear(input_dim, D_MODEL)

        # Positional encoding
        self.pos_enc = _PositionalEncoding(D_MODEL, max_len=50, dropout=TF_DROPOUT)

        # Transformer Encoder
        encoder_layer = nn.TransformerEncoderLayer(
            d_model=D_MODEL,
            nhead=NHEAD,
            dim_feedforward=TF_DIM_FF,
            dropout=TF_DROPOUT,
            batch_first=True,  # (batch, T, d_model)
            norm_first=True,  # Pre-LN: mais estável em datasets pequenos
        )
        self.encoder = nn.TransformerEncoder(encoder_layer, num_layers=TF_LAYERS)

        # Atenção temporal aprendível (mesmo design da GRU)
        self.attn = nn.Linear(D_MODEL, 1)

        # Cabeça de classificação (idêntica à GRU)
        self.head = nn.Sequential(
            nn.LayerNorm(D_MODEL),
            nn.Linear(D_MODEL, 32),
            nn.ReLU(),
            nn.Dropout(DROPOUT_HEAD),
            nn.Linear(32, 1),
        )

    def forward(self, x_num, x_cat):
        # Concatenar embeddings e features numéricas
        embs = [e(x_cat[..., i]) for i, e in enumerate(self.embeddings)]
        x = torch.cat([x_num] + embs, dim=-1) if embs else x_num  # (B, T, input_dim)

        # Projeção + positional encoding
        x = self.pos_enc(self.input_proj(x))  # (B, T, D_MODEL)

        # Transformer Encoder
        out = self.encoder(x)  # (B, T, D_MODEL)

        # Atenção temporal → contexto agregado
        w = torch.softmax(self.attn(out), dim=1)  # (B, T, 1)
        ctx = (w * out).sum(dim=1)  # (B, D_MODEL)

        return self.head(ctx).squeeze(1)  # (B,)


def run_transformer_seeds(all_records, cat_cols, num_cols, splits_by_seed, seeds):
    """M4_Transformer — Transformer Encoder com atenção temporal."""
    results = {}
    for seed in seeds:
        torch.manual_seed(seed)
        np.random.seed(seed)
        tr_ids, va_ids, te_ids = splits_by_seed[seed]
        tr = _filter_records(all_records, tr_ids)
        va = _filter_records(all_records, va_ids)
        te = _filter_records(all_records, te_ids)

        enc = _Encoder(cat_cols, num_cols)
        enc.fit(tr)

        g = torch.Generator()
        g.manual_seed(seed)
        _pin = DEVICE.type == "cuda"
        _nw = min(4, os.cpu_count() or 1) if _pin else 0
        mk = lambda recs, sh: DataLoader(
            _SeqDS(recs, enc),
            BATCH_SIZE,
            shuffle=sh,
            generator=g if sh else None,
            pin_memory=_pin,
            num_workers=_nw,
            persistent_workers=(_nw > 0),
        )
        tr_l, va_l, te_l = mk(tr, True), mk(va, False), mk(te, False)

        n_neg = sum(1 for r in tr if r["label"] == 0)
        n_pos = sum(1 for r in tr if r["label"] == 1)
        pos_w = torch.tensor([n_neg / n_pos], dtype=torch.float32).to(DEVICE)

        model = _Transformer(len(num_cols), enc.cat_vocab_sizes, enc.cat_emb_dims).to(
            DEVICE
        )
        opt = torch.optim.Adam(model.parameters(), lr=LR, weight_decay=WEIGHT_DECAY)
        crit = nn.BCEWithLogitsLoss(pos_weight=pos_w)
        sched = torch.optim.lr_scheduler.OneCycleLR(
            opt,
            max_lr=LR,
            steps_per_epoch=len(tr_l),
            epochs=EPOCHS,
            pct_start=0.1,
            anneal_strategy="cos",
        )

        best_loss, best_state, wait, epochs_run = float("inf"), None, 0, 0
        t0 = time.perf_counter()
        for _ in range(EPOCHS):
            epochs_run += 1
            _seq_train(model, tr_l, opt, crit, sched)
            vl, _, _, _ = _seq_eval(model, va_l, crit)
            if vl < best_loss:
                best_loss = vl
                best_state = {k: v.cpu().clone() for k, v in model.state_dict().items()}
                wait = 0
            else:
                wait += 1
                if wait >= PATIENCE:
                    break
        train_s = time.perf_counter() - t0

        model.load_state_dict(best_state)
        model.to(DEVICE)
        t1 = time.perf_counter()
        _, yp, yprob, yt = _seq_eval(model, te_l, crit)
        inf_s = time.perf_counter() - t1

        results[seed] = _result(
            yt,
            yp,
            yprob,
            len(tr),
            len(va),
            len(te),
            train_s,
            inf_s,
            epochs_run=epochs_run,
        )
    return results


# ══════════════════════════════════════════════════════════════════════════════
# TESTES ESTATÍSTICOS
# ══════════════════════════════════════════════════════════════════════════════

METHOD_ORDER = [
    "M1_LR_Snapshot",
    "M1_RF_Snapshot",
    "M1_CB_Snapshot",
    "M2_LR_VetorFlat",
    "M2_RF_VetorFlat",
    "M2_CB_VetorFlat",
    "M3_GRU",
    "M4_Transformer",
]


def run_statistical_tests(df_raw: pd.DataFrame) -> dict:
    """
    Para cada tempo T e cada métrica (accuracy, f1, auc):
      1. Teste de Friedman sobre os 6 métodos × 30 seeds
      2. Wilcoxon pareado para cada par de métodos (Bonferroni)

    Retorna dict:
      results[T][metric] = {
          "friedman": {"stat": float, "p": float},
          "wilcoxon": DataFrame(método_a × método_b → p_raw, p_bonf, significativo)
      }
    """
    out = {}
    metrics = ["accuracy", "f1", "auc"]

    for t in MINUTES_LIST:
        out[t] = {}
        sub_t = df_raw[df_raw["minute_cutoff"] == t].copy()

        for metric in metrics:
            # Matriz seeds × métodos
            mat = []
            valid_methods = []
            for m in METHOD_ORDER:
                col = sub_t[sub_t["method"] == m][metric].dropna().values
                if len(col) == N_SEEDS:
                    mat.append(col)
                    valid_methods.append(m)

            if len(mat) < 2:
                out[t][metric] = None
                continue

            # Friedman
            friedman_stat, friedman_p = friedmanchisquare(*mat)

            # Wilcoxon pareado com Bonferroni
            pairs = list(combinations(range(len(valid_methods)), 2))
            n_pairs = len(pairs)
            rows_wx = []
            for i, j in pairs:
                stat_wx, p_wx = wilcoxon(
                    mat[i], mat[j], alternative="two-sided", zero_method="wilcox"
                )
                p_bonf = min(p_wx * n_pairs, 1.0)
                rows_wx.append(
                    {
                        "metodo_a": valid_methods[i],
                        "metodo_b": valid_methods[j],
                        "stat_wx": round(stat_wx, 4),
                        "p_raw": round(p_wx, 6),
                        "p_bonferroni": round(p_bonf, 6),
                        "significativo": p_bonf < 0.05,
                    }
                )

            out[t][metric] = {
                "friedman": {
                    "stat": round(friedman_stat, 4),
                    "p": round(friedman_p, 6),
                    "methods": valid_methods,
                },
                "wilcoxon": pd.DataFrame(rows_wx),
            }

    return out


def save_wilcoxon_csvs(stat_results: dict):
    """Salva um CSV de Wilcoxon por (tempo, métrica)."""
    for t, metrics in stat_results.items():
        for metric, res in metrics.items():
            if res is None:
                continue
            path = OUTPUT_DIR / f"wilcoxon_T{t}_{metric}.csv"
            res["wilcoxon"].to_csv(path, index=False)
    print(f"[OK] CSVs Wilcoxon → {OUTPUT_DIR}/wilcoxon_*.csv")


# ══════════════════════════════════════════════════════════════════════════════
# ORQUESTRADOR
# ══════════════════════════════════════════════════════════════════════════════


def run_all():
    # Retomada
    if RAW_CSV.exists():
        df_done = pd.read_csv(RAW_CSV).dropna(subset=["accuracy"])
        done = set(
            zip(
                df_done["method"].astype(str),
                df_done["minute_cutoff"].astype(int),
                df_done["seed"].astype(int),
            )
        )
        rows = df_done.to_dict("records")
        print(f"[INFO] Retomando — {len(rows)} entradas já concluídas.")
    else:
        done, rows = set(), []

    def save():
        pd.DataFrame(rows).to_csv(RAW_CSV, index=False)

    total = 8 * len(MINUTES_LIST) * N_SEEDS
    done_n = len(rows)

    for minute_cutoff in MINUTES_LIST:
        print(f"\n{'═'*62}")
        print(f"  T = {minute_cutoff} min")
        print(f"{'═'*62}")

        # ── 1. matchIds comuns a todos os métodos ─────────────────────────────
        id_df = build_common_ids(DATASET_DIR, minute_cutoff)

        # ── 2. Splits por seed (uma vez, compartilhados por todos) ────────────
        splits_by_seed = {s: make_splits(id_df, s) for s in SEEDS}
        valid_ids = set(id_df["matchId"])

        # ── 3. Carrega representações UMA VEZ por tempo ───────────────────────
        print("  Carregando representações...")
        t_load = time.time()
        snap_df = load_snapshot(DATASET_DIR, minute_cutoff, valid_ids)
        flat_df = load_flat(DATASET_DIR, minute_cutoff, valid_ids)
        gru_data = load_gru(DATASET_DIR, minute_cutoff, valid_ids)
        print(
            f"  Carregamento: {time.time()-t_load:.1f}s  "
            f"(snap={len(snap_df)}, flat={len(flat_df)}, "
            f"gru={len(gru_data[0])} partidas)"
        )

        # ── 4. Métodos por tempo ───────────────────────────────────────────────
        methods_cfg = [
            # ── Snapshot ──────────────────────────────────────────────────────
            ("M1_LR_Snapshot", run_lr_snapshot_seeds, (snap_df,)),
            ("M1_RF_Snapshot", run_rf_snapshot_seeds, (snap_df,)),
            ("M1_CB_Snapshot", run_cb_snapshot_seeds, (snap_df,)),
            # ── Vetor Flat ────────────────────────────────────────────────────
            ("M2_LR_VetorFlat", run_lr_flat_seeds, (flat_df,)),
            ("M2_RF_VetorFlat", run_rf_flat_seeds, (flat_df,)),
            ("M2_CB_VetorFlat", run_cb_flat_seeds, (flat_df,)),
            # ── Sequencial ────────────────────────────────────────────────────
            ("M3_GRU", run_gru_seeds, gru_data),
            ("M4_Transformer", run_transformer_seeds, gru_data),
        ]

        for method_name, run_fn, data_args in methods_cfg:
            pending = [s for s in SEEDS if (method_name, minute_cutoff, s) not in done]
            if not pending:
                print(f"  [SKIP] {method_name} — já concluído.")
                continue

            print(f"\n  ── {method_name} ({len(pending)} seeds) ──")
            for seed in pending:
                done_n += 1
                print(f"    [{done_n}/{total}] seed={seed} ...", end=" ", flush=True)
                t0 = time.time()
                try:
                    res = run_fn(
                        *data_args, splits_by_seed=splits_by_seed, seeds=[seed]
                    )
                    r = res[seed]
                    elapsed = time.time() - t0
                    _auc = r.get("auc", float("nan"))
                    _tr = r.get("train_s", 0) or 0
                    _inf = r.get("inf_s", 0) or 0
                    print(
                        f"acc={r['accuracy']:.4f}  f1={r['f1']:.4f}"
                        + (
                            f"  auc={_auc:.4f}"
                            if not (isinstance(_auc, float) and _auc != _auc)
                            else ""
                        )
                        + f"  train={_tr:.1f}s  inf={_inf*1000:.1f}ms"
                    )
                    rows.append(
                        {
                            "method": method_name,
                            "minute_cutoff": minute_cutoff,
                            "seed": seed,
                            "accuracy": r["accuracy"],
                            "f1": r["f1"],
                            "auc": r.get("auc"),
                            "n_train": r.get("n_train"),
                            "n_val": r.get("n_val"),
                            "n_test": r.get("n_test"),
                            "epochs_run": r.get("epochs_run"),
                            "train_s": r.get("train_s"),
                            "inf_s": r.get("inf_s"),
                            "elapsed_s": round(elapsed, 2),
                        }
                    )
                    done.add((method_name, minute_cutoff, seed))
                except Exception as e:
                    print(f"ERRO: {e}")
                    rows.append(
                        {
                            "method": method_name,
                            "minute_cutoff": minute_cutoff,
                            "seed": seed,
                            "accuracy": None,
                            "f1": None,
                            "auc": None,
                            "n_train": None,
                            "n_val": None,
                            "n_test": None,
                            "epochs_run": None,
                            "train_s": None,
                            "inf_s": None,
                            "elapsed_s": None,
                        }
                    )
                save()

    print(f"\n[OK] Raw → {RAW_CSV}")
    return pd.DataFrame(rows)


# ══════════════════════════════════════════════════════════════════════════════
# PLANILHA
# ══════════════════════════════════════════════════════════════════════════════


def generate_xlsx(df_raw: pd.DataFrame, stat_results: dict):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    df = df_raw.dropna(subset=["accuracy", "f1"])
    wb = Workbook()

    H_FILL = PatternFill("solid", fgColor="1F4E79")
    A_FILL = PatternFill("solid", fgColor="D6E4F0")
    W_FILL = PatternFill("solid", fgColor="FFFFFF")
    S_FILL = PatternFill("solid", fgColor="C6EFCE")  # verde para sig.
    NS_FILL = PatternFill("solid", fgColor="FFCCCC")  # vermelho para não-sig.
    BOLD_W = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    BOLD_B = Font(bold=True, name="Arial", size=10)
    NORM = Font(name="Arial", size=10)
    CTR = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="AAAAAA")
    BRD = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hdr(cell, text):
        cell.value = text
        cell.font = BOLD_W
        cell.fill = H_FILL
        cell.alignment = CTR
        cell.border = BRD

    def val(cell, v, fmt=None, bold=False):
        cell.value = v
        cell.font = BOLD_B if bold else NORM
        cell.alignment = CTR
        cell.border = BRD
        if fmt:
            cell.number_format = fmt

    methods = METHOD_ORDER

    # ── Aba Resumo ────────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Resumo"
    RESUMO_COLS = [
        "Método",
        "T (min)",
        "N",
        "Acc Média",
        "Acc DP",
        "Acc IC95-",
        "Acc IC95+",
        "F1 Média",
        "F1 DP",
        "F1 IC95-",
        "F1 IC95+",
        "AUC Média",
        "AUC DP",
        "Train (s) médio",
        "Train (s) DP",
        "Inf (ms) médio",
        "Inf (ms) DP",
        "N Treino",
        "N Val",
        "N Teste",
    ]
    ws["A1"].value = (
        "Resultados: 30 Seeds × 4 Tempos × 8 Métodos — Splits Compartilhados"
    )
    ws["A1"].font = Font(bold=True, name="Arial", size=13, color="1F4E79")
    ws["A1"].alignment = CTR
    ws.row_dimensions[1].height = 28
    ws.merge_cells(f"A1:{get_column_letter(len(RESUMO_COLS))}1")
    for ci, h in enumerate(RESUMO_COLS, 1):
        hdr(ws.cell(2, ci), h)
    ws.row_dimensions[2].height = 36
    ri = 3
    toggle = False

    def _safe(v):
        return v if not (isinstance(v, float) and v != v) else None

    for method in methods:
        for t in MINUTES_LIST:
            sub = df[(df["method"] == method) & (df["minute_cutoff"] == t)]
            if sub.empty:
                continue
            n = len(sub)
            am, as_ = sub["accuracy"].mean(), sub["accuracy"].std(ddof=1)
            fm, fs = sub["f1"].mean(), sub["f1"].std(ddof=1)
            sub_auc = sub["auc"].dropna()
            auc_m = sub_auc.mean() if not sub_auc.empty else None
            auc_s = sub_auc.std(ddof=1) if len(sub_auc) > 1 else None
            tr_m = sub["train_s"].mean() if "train_s" in sub.columns else None
            tr_s = sub["train_s"].std(ddof=1) if "train_s" in sub.columns else None
            inf_m = sub["inf_s"].mean() * 1000 if "inf_s" in sub.columns else None
            inf_s2 = sub["inf_s"].std(ddof=1) * 1000 if "inf_s" in sub.columns else None
            nt_m = int(sub["n_train"].mean()) if "n_train" in sub.columns else None
            nv_m = int(sub["n_val"].mean()) if "n_val" in sub.columns else None
            nte_m = int(sub["n_test"].mean()) if "n_test" in sub.columns else None

            fill = A_FILL if toggle else W_FILL
            for c in range(1, len(RESUMO_COLS) + 1):
                ws.cell(ri, c).fill = fill

            val(ws.cell(ri, 1), method)
            val(ws.cell(ri, 2), t)
            val(ws.cell(ri, 3), n)
            val(ws.cell(ri, 4), am, "0.0000")
            val(ws.cell(ri, 5), as_, "0.0000")
            val(ws.cell(ri, 6), am - 1.96 * as_ / n**0.5, "0.0000")
            val(ws.cell(ri, 7), am + 1.96 * as_ / n**0.5, "0.0000")
            val(ws.cell(ri, 8), fm, "0.0000")
            val(ws.cell(ri, 9), fs, "0.0000")
            val(ws.cell(ri, 10), fm - 1.96 * fs / n**0.5, "0.0000")
            val(ws.cell(ri, 11), fm + 1.96 * fs / n**0.5, "0.0000")
            val(ws.cell(ri, 12), _safe(auc_m), "0.0000")
            val(ws.cell(ri, 13), _safe(auc_s), "0.0000")
            val(ws.cell(ri, 14), _safe(tr_m), "0.00")
            val(ws.cell(ri, 15), _safe(tr_s), "0.00")
            val(ws.cell(ri, 16), _safe(inf_m), "0.000")
            val(ws.cell(ri, 17), _safe(inf_s2), "0.000")
            val(ws.cell(ri, 18), nt_m)
            val(ws.cell(ri, 19), nv_m)
            val(ws.cell(ri, 20), nte_m)
            ri += 1
        toggle = not toggle

    col_w = [20, 8, 5, 11, 9, 11, 11, 11, 9, 11, 11, 11, 9, 13, 11, 13, 11, 9, 8, 8]
    for i, w in enumerate(col_w, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A3"

    # ── Aba Dados Brutos ──────────────────────────────────────────────────────
    wb2 = wb.create_sheet("Dados Brutos")
    cols_raw = [
        "method",
        "minute_cutoff",
        "seed",
        "accuracy",
        "f1",
        "auc",
        "n_train",
        "n_val",
        "n_test",
        "epochs_run",
        "train_s",
        "inf_s",
        "elapsed_s",
    ]
    for c in cols_raw:
        if c not in df.columns:
            df[c] = None
    FMT = {
        "accuracy": "0.0000",
        "f1": "0.0000",
        "auc": "0.0000",
        "train_s": "0.00",
        "elapsed_s": "0.00",
        "inf_s": "0.000000",
    }
    for ci, h in enumerate(cols_raw, 1):
        hdr(wb2.cell(1, ci), h)
    wb2.row_dimensions[1].height = 24
    for ri2, row in enumerate(df[cols_raw].itertuples(index=False), 2):
        for ci, v in enumerate(row, 1):
            c = wb2.cell(ri2, ci)
            c.value = v
            c.font = NORM
            c.alignment = CTR
            c.border = BRD
            fmt = FMT.get(cols_raw[ci - 1])
            if fmt:
                c.number_format = fmt
    for i, w in enumerate([20, 14, 8, 11, 11, 11, 9, 8, 8, 10, 11, 12, 11], 1):
        wb2.column_dimensions[get_column_letter(i)].width = w
    wb2.freeze_panes = "A2"

    # ── Aba Friedman ──────────────────────────────────────────────────────────
    wb4 = wb.create_sheet("Friedman")
    wb4["A1"].value = (
        "Teste de Friedman — p < 0.05 indica diferença global entre métodos"
    )
    wb4["A1"].font = Font(bold=True, name="Arial", size=11, color="1F4E79")
    wb4["A1"].alignment = CTR
    wb4.row_dimensions[1].height = 22
    wb4.merge_cells("A1:F1")
    for ci, h in enumerate(
        [
            "T (min)",
            "Métrica",
            "Estatística",
            "p-valor",
            "Significativo",
            "Métodos incluídos",
        ],
        1,
    ):
        hdr(wb4.cell(2, ci), h)
    ri4 = 3
    for t in MINUTES_LIST:
        for metric in ["accuracy", "f1", "auc"]:
            res = stat_results.get(t, {}).get(metric)
            if not res:
                continue
            fr = res["friedman"]
            sig = fr["p"] < 0.05
            cell_sig = wb4.cell(ri4, 5)
            val(wb4.cell(ri4, 1), t)
            val(wb4.cell(ri4, 2), metric)
            val(wb4.cell(ri4, 3), fr["stat"], "0.0000")
            val(wb4.cell(ri4, 4), fr["p"], "0.000000")
            cell_sig.value = "Sim" if sig else "Não"
            cell_sig.font = BOLD_B
            cell_sig.alignment = CTR
            cell_sig.border = BRD
            cell_sig.fill = S_FILL if sig else NS_FILL
            val(wb4.cell(ri4, 6), ", ".join(fr["methods"]))
            ri4 += 1
    for i, w in enumerate([10, 12, 14, 14, 14, 60], 1):
        wb4.column_dimensions[get_column_letter(i)].width = w
    wb4.freeze_panes = "A3"

    # ── Aba Wilcoxon ──────────────────────────────────────────────────────────
    for t in MINUTES_LIST:
        ws_wx = wb.create_sheet(f"Wilcoxon_T{t}")
        ws_wx["A1"].value = (
            f"Wilcoxon pareado (Bonferroni) — T={t} min  |  "
            f"Significativo = p_bonf < 0.05"
        )
        ws_wx["A1"].font = Font(bold=True, name="Arial", size=11, color="1F4E79")
        ws_wx["A1"].alignment = CTR
        ws_wx.row_dimensions[1].height = 22
        ws_wx.merge_cells("A1:H1")

        ri_wx = 2
        for metric in ["accuracy", "f1", "auc"]:
            res = stat_results.get(t, {}).get(metric)
            if not res:
                continue
            hdr(ws_wx.cell(ri_wx, 1), f"Métrica: {metric}")
            ri_wx += 1
            for ci, h in enumerate(
                [
                    "Método A",
                    "Método B",
                    "Stat Wilcoxon",
                    "p raw",
                    "p Bonferroni",
                    "Sig. (α=0.05)",
                    "Melhor",
                ],
                1,
            ):
                hdr(ws_wx.cell(ri_wx, ci), h)
            ri_wx += 1

            for _, row_wx in res["wilcoxon"].iterrows():
                sig = row_wx["significativo"]
                # Identifica qual método tem maior média nesta métrica/tempo
                m_a = df[
                    (df["method"] == row_wx["metodo_a"]) & (df["minute_cutoff"] == t)
                ][metric].mean()
                m_b = df[
                    (df["method"] == row_wx["metodo_b"]) & (df["minute_cutoff"] == t)
                ][metric].mean()
                melhor = (
                    (row_wx["metodo_a"] if m_a >= m_b else row_wx["metodo_b"])
                    if sig
                    else "—"
                )
                for ci, v in enumerate(
                    [
                        row_wx["metodo_a"],
                        row_wx["metodo_b"],
                        row_wx["stat_wx"],
                        row_wx["p_raw"],
                        row_wx["p_bonferroni"],
                        "Sim" if sig else "Não",
                        melhor,
                    ],
                    1,
                ):
                    c = ws_wx.cell(ri_wx, ci)
                    c.value = v
                    c.font = NORM
                    c.alignment = CTR
                    c.border = BRD
                    if ci == 6:
                        c.fill = S_FILL if sig else NS_FILL
                        c.font = BOLD_B
                    if ci in (3, 4, 5):
                        c.number_format = "0.000000"
                ri_wx += 1
            ri_wx += 1  # linha em branco entre métricas

        for i, w in enumerate([22, 22, 16, 14, 14, 14, 22], 1):
            ws_wx.column_dimensions[get_column_letter(i)].width = w
        ws_wx.freeze_panes = "A3"

    wb.save(SUMMARY_XLSX)
    print(f"[OK] Planilha → {SUMMARY_XLSX}")


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    print("=" * 62)
    print(f"  Experimentos: {N_SEEDS} seeds × {len(MINUTES_LIST)} tempos × 8 métodos")
    print(f"  Splits COMPARTILHADOS por tempo (válido para testes pareados)")
    print(f"  Device : {DEVICE}")
    print(f"  Tempos : {MINUTES_LIST}")
    print(f"  Total  : {8 * len(MINUTES_LIST) * N_SEEDS} rodadas")
    print("=" * 62)

    df_raw = run_all()
    stat_results = run_statistical_tests(df_raw)
    save_wilcoxon_csvs(stat_results)
    generate_xlsx(df_raw, stat_results)

    print("\n[CONCLUÍDO]")
    print(f"  {RAW_CSV}")
    print(f"  {SUMMARY_XLSX}")
    print(f"  {OUTPUT_DIR}/wilcoxon_*.csv")
